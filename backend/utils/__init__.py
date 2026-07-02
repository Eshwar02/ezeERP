"""Utilities: invoice PDF generation (ReportLab), report Excel export (openpyxl), Appwrite storage."""
import io
import uuid

from appwrite.client import Client
from appwrite.services.storage import Storage
from appwrite.input_file import InputFile
from config import Config

def _appwrite_storage() -> Storage:
    client = Client()
    client.set_endpoint(Config.APPWRITE_ENDPOINT)
    client.set_project(Config.APPWRITE_PROJECT)
    client.set_key(Config.APPWRITE_KEY)
    return Storage(client)

def upload_file(buf: io.BytesIO, filename: str, mime: str) -> dict:
    """Upload buf to Appwrite bucket. Returns Appwrite file document."""
    storage = _appwrite_storage()
    file_id = uuid.uuid4().hex
    result = storage.create_file(
        bucket_id=Config.APPWRITE_BUCKET,
        file_id=file_id,
        file=InputFile.from_bytes(buf.read(), filename=filename, mime_type=mime),
    )
    return result

def get_file_url(file_id: str) -> str:
    """Return direct download URL for a stored file."""
    return (
        f"{Config.APPWRITE_ENDPOINT}/storage/buckets/{Config.APPWRITE_BUCKET}"
        f"/files/{file_id}/download"
        f"?project={Config.APPWRITE_PROJECT}"
    )


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# ezeERP brand palette: green / white / blue
BRAND_GREEN = colors.HexColor("#21c25e")
BRAND_BLUE = colors.HexColor("#00a8ec")


def build_invoice_pdf(company, customer, sale):
    """Return a BytesIO containing the invoice PDF for a sale."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    title = styles["Title"]
    title.textColor = BRAND_BLUE
    normal = styles["Normal"]

    elements = [
        Paragraph(company.name if company else "ezeERP", title),
        Paragraph("Tax Invoice", styles["Heading2"]),
    ]
    meta_bits = []
    if company:
        if company.address:
            meta_bits.append(company.address)
        if company.gst_number:
            meta_bits.append(f"GSTIN: {company.gst_number}")
    if meta_bits:
        elements.append(Paragraph(" &nbsp;|&nbsp; ".join(meta_bits), normal))
    elements.append(Spacer(1, 8))

    info = [
        ["Invoice #", sale.invoice_number, "Date", sale.date.strftime("%Y-%m-%d") if sale.date else ""],
        ["Bill To", customer.name if customer else "Cash", "Mobile", (customer.mobile if customer else "") or ""],
    ]
    info_tbl = Table(info, colWidths=[28 * mm, 70 * mm, 24 * mm, 50 * mm])
    info_tbl.setStyle(
        TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TEXTCOLOR", (0, 0), (0, -1), BRAND_GREEN),
            ("TEXTCOLOR", (2, 0), (2, -1), BRAND_GREEN),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])
    )
    elements.extend([info_tbl, Spacer(1, 10)])

    head = ["#", "Item", "Qty", "Rate", "GST%", "Amount"]
    rows = [head]
    for i, it in enumerate(sale.items, 1):
        rows.append([
            str(i), it.name, f"{it.quantity:g}", f"{it.rate:.2f}",
            f"{it.gst_percentage:g}", f"{it.amount:.2f}",
        ])
    rows.append(["", "", "", "", "Subtotal", f"{sale.subtotal:.2f}"])
    rows.append(["", "", "", "", "Tax", f"{sale.tax_total:.2f}"])
    rows.append(["", "", "", "", "Total", f"{sale.total:.2f}"])

    tbl = Table(rows, colWidths=[10 * mm, 70 * mm, 18 * mm, 24 * mm, 18 * mm, 28 * mm])
    tbl.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BRAND_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -4), 0.4, colors.HexColor("#d0d0d0")),
            ("LINEABOVE", (4, -3), (-1, -3), 0.5, BRAND_GREEN),
            ("FONTNAME", (4, -1), (-1, -1), "Helvetica-Bold"),
            ("TEXTCOLOR", (4, -1), (-1, -1), BRAND_GREEN),
            ("ROWBACKGROUNDS", (0, 1), (-1, -4), [colors.white, colors.HexColor("#f3fbf6")]),
        ])
    )
    elements.append(tbl)
    elements.append(Spacer(1, 16))
    elements.append(Paragraph("Thank you for your business — powered by ezeERP", normal))

    doc.build(elements)
    buf.seek(0)
    return buf


_XLSX_HEAD_FILL = PatternFill("solid", fgColor="21C25E")
_XLSX_HEAD_FONT = Font(bold=True, color="FFFFFF")


def build_report_xlsx(title, sections):
    """Return a BytesIO .xlsx for a report.

    sections: list of {name, headers:[str], rows:[[cell]], total?:[cell]}.
    Each section becomes a heading row, a styled header row, data rows, and an
    optional bold total row.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    r = 1
    ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=14, color="00A8EC")
    r += 2

    for sec in sections:
        if sec.get("name"):
            ws.cell(row=r, column=1, value=sec["name"]).font = Font(bold=True, size=12)
            r += 1
        headers = sec.get("headers") or []
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.fill = _XLSX_HEAD_FILL
            cell.font = _XLSX_HEAD_FONT
        r += 1
        for row in sec.get("rows") or []:
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
            r += 1
        total = sec.get("total")
        if total:
            for c, val in enumerate(total, 1):
                ws.cell(row=r, column=c, value=val).font = Font(bold=True)
            r += 1
        r += 1

    for col in ws.columns:
        width = max((len(str(cell.value)) for cell in col if cell.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
